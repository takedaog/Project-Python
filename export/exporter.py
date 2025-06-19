#Реализация функции export_to_xlsx() 

from openpyxl import Workbook
from data.database import users, assignments, grades, schedules, notifications

def export_to_xlsx(filename="eduplatform_export.xlsx"):
    """Экспортирует данные из памяти в .xlsx файл"""

    wb = Workbook()

    #Users sheet
    ws_users = wb.active
    ws_users.title = "Users"
    ws_users.append(["ID", "Name", "Email", "Role", "Created At"])
    for user in users.values():
        profile = user.get_profile()
        ws_users.append([
            profile['id'],
            profile['name'],
            profile['email'],
            profile['role'],
            profile['created_at']
        ])

    #Assignments sheet 
    ws_assignments = wb.create_sheet("Assignments")
    ws_assignments.append(["ID", "Title", "Subject", "Class", "Teacher ID", "Deadline", "Submissions", "Grades"])
    for assignment in assignments.values():
        ws_assignments.append([
            assignment.id,
            assignment.title,
            assignment.subject,
            assignment.class_id,
            assignment.teacher_id,
            assignment.deadline,
            str(assignment.submissions),
            str(assignment.grades)
        ])

    # Grades sheet
    ws_grades = wb.create_sheet("Grades")
    ws_grades.append(["ID", "Student ID", "Subject", "Value", "Date", "Teacher ID", "Comment"])
    for grade in grades:
        g = grade.get_grade_info()
        ws_grades.append([
            g["id"],
            g["student_id"],
            g["subject"],
            g["value"],
            g["date"],
            g["teacher_id"],
            g["comment"]
        ])

    #  Schedules sheet
    ws_schedules = wb.create_sheet("Schedules")
    ws_schedules.append(["ID", "Class", "Day", "Lessons"])
    for sched in schedules.values():
        ws_schedules.append([
            sched.id,
            sched.class_id,
            sched.day,
            str(sched.lessons)
        ])

    #Notifications sheet 
    ws_notifications = wb.create_sheet("Notifications")
    ws_notifications.append(["ID", "Message", "Recipient", "Created At", "Is Read", "Priority"])
    for notif in notifications:
        ws_notifications.append([
            notif["id"],
            notif["message"],
            notif["recipient_id"],
            notif["created_at"],
            notif["is_read"],
            notif["priority"]
        ])

    # Save workbook
    wb.save(filename)
    print(f"✅ Exported to {filename}")


#export_to_csv() для экспорта всех таблиц в отдельные CSV-файлы

import csv
from data.database import users, assignments, grades, schedules, notifications

def export_to_csv():
    """📄 Экспортирует данные в отдельные CSV-файлы по таблицам"""

    # Users.csv 
    with open("Users.csv", mode="w", newline='', encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["ID", "Name", "Email", "Role", "Created At"])
        for user in users.values():
            profile = user.get_profile()
            writer.writerow([
                profile["id"],
                profile["name"],
                profile["email"],
                profile["role"],
                profile["created_at"]
            ])

    # Assignments.csv
    with open("Assignments.csv", mode="w", newline='', encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["ID", "Title", "Subject", "Class", "Teacher ID", "Deadline", "Submissions", "Grades"])
        for a in assignments.values():
            writer.writerow([
                a.id,
                a.title,
                a.subject,
                a.class_id,
                a.teacher_id,
                a.deadline,
                str(a.submissions),
                str(a.grades)
            ])

    # Grades.csv
    with open("Grades.csv", mode="w", newline='', encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["ID", "Student ID", "Subject", "Value", "Date", "Teacher ID", "Comment"])
        for g in grades:
            info = g.get_grade_info()
            writer.writerow([
                info["id"],
                info["student_id"],
                info["subject"],
                info["value"],
                info["date"],
                info["teacher_id"],
                info["comment"]
            ])

    # Schedules.csv
    with open("Schedules.csv", mode="w", newline='', encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["ID", "Class", "Day", "Lessons"])
        for s in schedules.values():
            writer.writerow([
                s.id,
                s.class_id,
                s.day,
                str(s.lessons)
            ])

    #Notifications.csv
    with open("Notifications.csv", mode="w", newline='', encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["ID", "Message", "Recipient ID", "Created At", "Is Read", "Priority"])
        for n in notifications:
            writer.writerow([
                n["id"],
                n["message"],
                n["recipient_id"],
                n["created_at"],
                n["is_read"],
                n["priority"]
            ])

    print("All CSV files exported successfully.")



#Экспорт всех таблиц в SQL-совместимый .sql файл

from data.database import users, assignments, grades, schedules, notifications

def export_to_sql(filename="eduplatform_export.sql"):
    """📦 Экспортирует данные в SQL-совместимый файл"""

    sql_lines = []

    # --- CREATE TABLES ---

    sql_lines.append("""
-- Users
CREATE TABLE Users (
    id INT PRIMARY KEY,
    full_name VARCHAR(255),
    email VARCHAR(255),
    role VARCHAR(50),
    created_at VARCHAR(50)
);
""")

    sql_lines.append("""
-- Assignments
CREATE TABLE Assignments (
    id INT PRIMARY KEY,
    title VARCHAR(255),
    subject VARCHAR(100),
    class_id VARCHAR(20),
    teacher_id INT,
    deadline VARCHAR(50),
    submissions TEXT,
    grades TEXT
);
""")

    sql_lines.append("""
-- Grades
CREATE TABLE Grades (
    id INT PRIMARY KEY,
    student_id INT,
    subject VARCHAR(100),
    value INT CHECK (value >= 1 AND value <= 5),
    date VARCHAR(50),
    teacher_id INT,
    comment TEXT
);
""")

    sql_lines.append("""
-- Schedules
CREATE TABLE Schedules (
    id INT PRIMARY KEY,
    class_id VARCHAR(20),
    day VARCHAR(20),
    lessons TEXT
);
""")

    sql_lines.append("""
-- Notifications
CREATE TABLE Notifications (
    id INT PRIMARY KEY,
    message TEXT,
    recipient_id INT,
    created_at VARCHAR(50),
    is_read BIT,
    priority VARCHAR(20)
);
""")

    # --- INSERT DATA ---

    for u in users.values():
        p = u.get_profile()
        sql_lines.append(f"INSERT INTO Users VALUES ({p['id']}, '{p['name']}', '{p['email']}', '{p['role']}', '{p['created_at']}');")

    for a in assignments.values():
        sql_lines.append(f"INSERT INTO Assignments VALUES ({a.id}, '{a.title}', '{a.subject}', '{a.class_id}', {a.teacher_id}, '{a.deadline}', '{str(a.submissions)}', '{str(a.grades)}');")

    for g in grades:
        g_info = g.get_grade_info()
        sql_lines.append(
            f"INSERT INTO Grades VALUES ({g_info['id']}, {g_info['student_id']}, '{g_info['subject']}', {g_info['value']}, '{g_info['date']}', {g_info['teacher_id']}, '{g_info['comment']}');"
        )

    for s in schedules.values():
        sql_lines.append(f"INSERT INTO Schedules VALUES ({s.id}, '{s.class_id}', '{s.day}', '{str(s.lessons)}');")

    for n in notifications:
        is_read = 1 if n["is_read"] else 0
        sql_lines.append(f"INSERT INTO Notifications VALUES ({n['id']}, '{n['message']}', {n['recipient_id']}, '{n['created_at']}', {is_read}, '{n['priority']}');")

    # --- Save to file ---
    with open(filename, "w", encoding="utf-8") as f:
        f.write("\n".join(sql_lines))

    print(f"SQL export saved to {filename}")
